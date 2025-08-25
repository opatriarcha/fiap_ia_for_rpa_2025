import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def gerar_dados():
    """Cria dois DataFrames com dados fictícios de pessoas físicas e empresas"""
    pessoas = {
        'Nome': [f'Pessoa {i}' for i in range(1, 101)],
        'CPF': [f'000.000.000-{i:02d}' for i in range(1, 101)],
        'Idade': [20 + i % 40 for i in range(100)],
        'Email': [f'pessoa{i}@exemplo.com' for i in range(1, 101)],
        'Cidade': [f'Cidade {i % 10}' for i in range(100)],
    }

    empresas = {
        'Razão Social': [f'Empresa {i}' for i in range(1, 101)],
        'CNPJ': [f'00.000.000/000{i:02d}' for i in range(1, 101)],
        'Segmento': [f'Segmento {i % 5}' for i in range(100)],
        'Telefone': [f'(11) 4002-89{i % 10}' for i in range(100)],
        'Estado': [f'Estado {i % 5}' for i in range(100)],
    }

    return pd.DataFrame(pessoas), pd.DataFrame(empresas)

def formatar_cabecalho(sheet):
    """Aplica formatação visual ao cabeçalho da planilha"""
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4F81BD")
    align = Alignment(horizontal="center")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = align

def adicionar_tabela_com_formula(sheet, nome_tabela):
    """Adiciona uma tabela formatada e insere fórmula de contagem"""
    total_linhas = sheet.max_row
    total_colunas = sheet.max_column
    ultima_coluna_letra = chr(64 + total_colunas)
    ref = f"A1:{ultima_coluna_letra}{total_linhas}"

    tabela = Table(displayName=nome_tabela, ref=ref)
    estilo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tabela.tableStyleInfo = estilo
    sheet.add_table(tabela)

    # Fórmula para contar registros
    sheet[f"{ultima_coluna_letra}{total_linhas + 2}"] = "Total de registros:"
    sheet[f"{ultima_coluna_letra}{total_linhas + 3}"] = f"=ROWS({ref})-1"

def criar_planilha_formatada(path_saida="dados_unificados_formatados.xlsx"):
    df_pf, df_pj = gerar_dados()

    wb = Workbook()
    ws_pf = wb.active
    ws_pf.title = "Pessoas Físicas"
    ws_pj = wb.create_sheet("Empresas")

    # Inserir dados nas planilhas
    for linha in dataframe_to_rows(df_pf, index=False, header=True):
        ws_pf.append(linha)
    for linha in dataframe_to_rows(df_pj, index=False, header=True):
        ws_pj.append(linha)

    # Aplicar formatação
    formatar_cabecalho(ws_pf)
    formatar_cabecalho(ws_pj)

    adicionar_tabela_com_formula(ws_pf, "TabelaPF")
    adicionar_tabela_com_formula(ws_pj, "TabelaPJ")

    # Salvar
    wb.save(path_saida)
    print(f"Arquivo salvo com sucesso em: {path_saida}")

# Execução
if __name__ == "__main__":
    criar_planilha_formatada()
